# File: find_and_list_genes.py
# Author: Dennis Kovarik
# Purpose: To search the RAST and PATRIC genome annotations for genes
#          containing keywords
#
# Description: This program parses the RAST and Genome excel files for genes
# containing keywords. The program will prompt the user to enter keywords that
# will be used to search for matching gene names. If multiple keywords are 
# entered per entry, the program will search for and list out any gene names
# that contain all of those keywords in any order.
#
# Usage:
#   find_and_list_genes.py path\to\RAST_file.xls path\to\PATRIC_file.xlsx

# Import Modules
import sys
import os
import csv
import codecs
import pandas as pd
import openpyxl
from utils import *


def read_PATRIC_genes(filepath):
    """
    This function reads all the genes in an excel spreadsheet from a 
    genome annotation completed by PATRIC. This function will store these  
    genes in a dictionary of genes and their counts and return it.
    
    :param filepath: The filepath of the PATRIC file
    :return: A dictionary of genes and their counts
    """
    if not is_PATRIC_spreadsheet(filepath):
        raise Exception("File is not a RAST excel spreadsheet")
    genes = {} 
    try:       
        wb = openpyxl.load_workbook(filepath)
        first_sheet = wb.sheetnames[0]
        worksheet = wb[first_sheet]
        # here you iterate over the rows in the specific column
        for row in range(2,worksheet.max_row+1):  
            for column in "T":  
                cell_name = "{}{}".format(column, row)
                gene = worksheet[cell_name].value
                if gene in genes.keys(): 
                    genes[gene] += 1
                else:
                    genes[gene.strip()] = 1       
        return genes
    except:
        err = "An error occured while reading file column names for PATRIC "
        err += "excel file."
        print(err)
    return genes


def read_RAST_genes(filepath):
    """
    This function reads all the genes in an excel spreadsheet from a 
    genome annotation completed by RAST. This function will store these genes 
    in a dictionary and return it.
    
    :param filepath: The filepath of the RAST file
    :return: A dictionary of proteins with the duplicate counts
    """
    if not is_RAST_spreadsheet(filepath):
        raise Exception("File is not a RAST excel spreadsheet")
        
    # Read the xls file
    df = pd.read_excel(filepath)
    # Store EC numbers in a python set
    genes = {} 
    for gene in df.function:
        if gene in genes.keys(): 
            genes[gene] += 1
        else:
            genes[gene.strip()] = 1
    return genes



# Check the number of command line arguments
if len(sys.argv) < 3:
    print("Invalid Number of Command Line Arguments")
    exit()
    
    
# Read RAST Genes
genes_RAST = read_RAST_genes(sys.argv[1])
genes_PATRIC = read_PATRIC_genes(sys.argv[2])


while True:
    key = input("Enter search term: ")
    key = key.lower()
    out = []
    num_genes = 0
    for gene in genes_RAST.keys():
        good = True
        for word in key.split(" "):
            if gene.lower().find(word) == -1:
                good = False
        if good:
            num_genes += genes_RAST[gene]
            out.append(gene + " (" + str(genes_RAST[gene]) + ")")
           
    print("RAST")
    print("-------------------------------------------------------------")
    print("Number of Genes: " + str(num_genes))
    print("-------------------------------------------------------------")
    for g in out:
        print(g)
            
    out = []
    num_genes = 0
    for gene in genes_PATRIC.keys():
        good = True
        for word in key.split(" "):
            if gene.lower().find(word) == -1:
                good = False
        if good:
            num_genes += genes_PATRIC[gene]
            out.append(gene + " (" + str(genes_PATRIC[gene]) + ")")
            
    print('\n')
    print("PATRIC")
    print("-------------------------------------------------------------")
    print("Number of Genes: " + str(num_genes))
    print("-------------------------------------------------------------")
    for g in out:
        print(g)        
    print("\nKeywords: " + key)
            
            
    print("\n\n")
            
            